import logging
from io import BytesIO
from pathlib import Path
from typing import Annotated, Any, Optional, Union, cast

from docling_core.types.doc import (
    BoundingBox,
    ContentLayer,
    CoordOrigin,
    DocItem,
    DoclingDocument,
    DocumentOrigin,
    GroupLabel,
    ImageRef,
    ProvenanceItem,
    Size,
    TableCell,
    TableData,
)
from openpyxl import load_workbook
from openpyxl.chartsheet.chartsheet import Chartsheet
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor
from openpyxl.worksheet.worksheet import Worksheet
from PIL import Image as PILImage
from pydantic import BaseModel, Field, NonNegativeInt, PositiveInt
from pydantic.dataclasses import dataclass
from typing_extensions import override

from docling.backend.abstract_backend import (
    DeclarativeDocumentBackend,
    PaginatedDocumentBackend,
)
from docling.datamodel.base_models import InputFormat
from docling.datamodel.document import InputDocument

_log = logging.getLogger(__name__)


@dataclass
class DataRegion:
    """Represents the bounding rectangle of non-empty cells in a worksheet."""

    min_row: Annotated[
        PositiveInt, Field(description="Smallest row index (1-based index).")
    ]
    max_row: Annotated[
        PositiveInt, Field(description="Largest row index (1-based index).")
    ]
    min_col: Annotated[
        PositiveInt, Field(description="Smallest column index (1-based index).")
    ]
    max_col: Annotated[
        PositiveInt, Field(description="Largest column index (1-based index).")
    ]

    def width(self) -> PositiveInt:
        """Number of columns in the data region."""
        return self.max_col - self.min_col + 1

    def height(self) -> PositiveInt:
        """Number of rows in the data region."""
        return self.max_row - self.min_row + 1


class ExcelCell(BaseModel):
    """Represents an Excel cell.

    Attributes:
        row: The row number of the cell.
        col: The column number of the cell.
        text: The text content of the cell.
        row_span: The number of rows the cell spans.
        col_span: The number of columns the cell spans.
    """

    row: int
    col: int
    text: str
    row_span: int
    col_span: int


class ExcelTable(BaseModel):
    """Represents an Excel table on a worksheet.

    Attributes:
        anchor: The column and row indices of the upper-left cell of the table
        (0-based index).
        num_rows: The number of rows in the table.
        num_cols: The number of columns in the table.
        data: The data in the table, represented as a list of ExcelCell objects.
    """

    anchor: tuple[NonNegativeInt, NonNegativeInt]
    num_rows: int
    num_cols: int
    data: list[ExcelCell]


class MsExcelDocumentBackend(DeclarativeDocumentBackend, PaginatedDocumentBackend):
    """Backend for parsing Excel workbooks.

    The backend converts an Excel workbook into a DoclingDocument object.
    Each worksheet is converted into a separate page.
    The following elements are parsed:
    - Cell contents, parsed as tables. If two groups of cells are disconnected
    between each other, they will be parsed as two different tables.
    - Images, parsed as PictureItem objects.

    The DoclingDocument tables and pictures have their provenance information, including
    the position in their original Excel worksheet. The position is represented by a
    bounding box object with the cell indices as units (0-based index). The size of this
    bounding box is the number of columns and rows that the table or picture spans.
    """

    @override
    def __init__(
        self, in_doc: "InputDocument", path_or_stream: Union[BytesIO, Path]
    ) -> None:
        """Initialize the MsExcelDocumentBackend object.

        Parameters:
            in_doc: The input document object.
            path_or_stream: The path or stream to the Excel file.

        Raises:
            RuntimeError: An error occurred parsing the file.
        """
        super().__init__(in_doc, path_or_stream)

        # Initialise the parents for the hierarchy
        self.max_levels = 10

        self.parents: dict[int, Any] = {}
        for i in range(-1, self.max_levels):
            self.parents[i] = None

        self.workbook = None
        try:
            if isinstance(self.path_or_stream, BytesIO):
                self.workbook = load_workbook(
                    filename=self.path_or_stream, data_only=True
                )

            elif isinstance(self.path_or_stream, Path):
                self.workbook = load_workbook(
                    filename=str(self.path_or_stream), data_only=True
                )

            self.valid = self.workbook is not None
        except Exception as e:
            self.valid = False

            raise RuntimeError(
                f"MsExcelDocumentBackend could not load document with hash {self.document_hash}"
            ) from e

    @override
    def is_valid(self) -> bool:
        _log.debug(f"valid: {self.valid}")
        return self.valid

    @classmethod
    @override
    def supports_pagination(cls) -> bool:
        return True

    @override
    def page_count(self) -> int:
        if self.is_valid() and self.workbook:
            return len(self.workbook.sheetnames)
        else:
            return 0

    @classmethod
    @override
    def supported_formats(cls) -> set[InputFormat]:
        return {InputFormat.XLSX}

    @override
    def convert(self) -> DoclingDocument:
        """Parse the Excel workbook into a DoclingDocument object.

        Raises:
            RuntimeError: Unable to run the conversion since the backend object failed to
            initialize.

        Returns:
            The DoclingDocument object representing the Excel workbook.
        """
        origin = DocumentOrigin(
            filename=self.file.name or "file.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            binary_hash=self.document_hash,
        )

        doc = DoclingDocument(name=self.file.stem or "file.xlsx", origin=origin)

        if self.is_valid():
            doc = self._convert_workbook(doc)
        else:
            raise RuntimeError(
                f"Cannot convert doc with {self.document_hash} because the backend failed to init."
            )

        return doc

    def _convert_workbook(self, doc: DoclingDocument) -> DoclingDocument:
        """Parse the Excel workbook and attach its structure to a DoclingDocument.

        Args:
            doc: A DoclingDocument object.

        Returns:
            A DoclingDocument object with the parsed items.
        """

        if self.workbook is not None:
            # Iterate over all sheets
            for idx, name in enumerate(self.workbook.sheetnames):
                _log.info(f"Processing sheet {idx}: {name}")

                sheet = self.workbook[name]
                page_no = idx + 1
                # do not rely on sheet.max_column, sheet.max_row if there are images
                page = doc.add_page(page_no=page_no, size=Size(width=0, height=0))

                self.parents[0] = doc.add_group(
                    parent=None,
                    label=GroupLabel.SECTION,
                    name=f"sheet: {name}",
                    content_layer=self._get_sheet_content_layer(sheet),
                )
                doc = self._convert_sheet(doc, sheet)
                width, height = self._find_page_size(doc, page_no)
                page.size = Size(width=width, height=height)
        else:
            _log.error("Workbook is not initialized.")

        return doc

    def _convert_sheet(
        self, doc: DoclingDocument, sheet: Union[Worksheet, Chartsheet]
    ) -> DoclingDocument:
        """Parse an Excel worksheet and attach its structure to a DoclingDocument

        Args:
            doc: The DoclingDocument to be updated.
            sheet: The Excel worksheet to be parsed.

        Returns:
            The updated DoclingDocument.
        """
        if isinstance(sheet, Worksheet):
            doc = self._find_tables_in_sheet(doc, sheet)
            doc = self._find_images_in_sheet(doc, sheet)

        # TODO: parse charts in sheet

        return doc

    def _find_tables_in_sheet(
        self, doc: DoclingDocument, sheet: Worksheet
    ) -> DoclingDocument:
        """Find all tables in an Excel sheet and attach them to a DoclingDocument.

        Args:
            doc: The DoclingDocument to be updated.
            sheet: The Excel worksheet to be parsed.

        Returns:
            The updated DoclingDocument.
        """

        if self.workbook is not None:
            content_layer = self._get_sheet_content_layer(sheet)
            tables = self._find_data_tables(sheet)

            for excel_table in tables:
                origin_col = excel_table.anchor[0]
                origin_row = excel_table.anchor[1]
                num_rows = excel_table.num_rows
                num_cols = excel_table.num_cols

                table_data = TableData(
                    num_rows=num_rows,
                    num_cols=num_cols,
                    table_cells=[],
                )

                for excel_cell in excel_table.data:
                    cell = TableCell(
                        text=excel_cell.text,
                        row_span=excel_cell.row_span,
                        col_span=excel_cell.col_span,
                        start_row_offset_idx=excel_cell.row,
                        end_row_offset_idx=excel_cell.row + excel_cell.row_span,
                        start_col_offset_idx=excel_cell.col,
                        end_col_offset_idx=excel_cell.col + excel_cell.col_span,
                        column_header=excel_cell.row == 0,
                        row_header=False,
                    )
                    table_data.table_cells.append(cell)

                page_no = self.workbook.index(sheet) + 1
                doc.add_table(
                    data=table_data,
                    parent=self.parents[0],
                    prov=ProvenanceItem(
                        page_no=page_no,
                        charspan=(0, 0),
                        bbox=BoundingBox.from_tuple(
                            (
                                origin_col,
                                origin_row,
                                origin_col + num_cols,
                                origin_row + num_rows,
                            ),
                            origin=CoordOrigin.TOPLEFT,
                        ),
                    ),
                    content_layer=content_layer,
                )

        return doc

    def _find_true_data_bounds(self, sheet: Worksheet) -> DataRegion:
        """Find the true data boundaries (min/max rows and columns) in a worksheet.

        This function scans all cells to find the smallest rectangular region that contains
        all non-empty cells or merged cell ranges. It returns the minimal and maximal
        row/column indices that bound the actual data region.

        Args:
            sheet: The worksheet to analyze.

        Returns:
            A data region representing the smallest rectangle that covers all data and merged cells.
            If the sheet is empty, returns (1, 1, 1, 1) by default.
        """
        min_row, min_col = None, None
        max_row, max_col = 0, 0

        for cell in sheet._cells.values():
            if cell.value is not None:
                r, c = cell.row, cell.column
                min_row = r if min_row is None else min(min_row, r)
                min_col = c if min_col is None else min(min_col, c)
                max_row = max(max_row, r)
                max_col = max(max_col, c)

        # Expand bounds to include merged cells
        for merged in sheet.merged_cells.ranges:
            min_row = (
                merged.min_row if min_row is None else min(min_row, merged.min_row)
            )
            min_col = (
                merged.min_col if min_col is None else min(min_col, merged.min_col)
            )
            max_row = max(max_row, merged.max_row)
            max_col = max(max_col, merged.max_col)

        # If no data found, default to (1, 1, 1, 1)
        if min_row is None or min_col is None:
            min_row = min_col = max_row = max_col = 1

        return DataRegion(min_row, max_row, min_col, max_col)

    def _find_data_tables(self, sheet: Worksheet) -> list[ExcelTable]:
        """Find all compact rectangular data tables in an Excel worksheet.

        Args:
            sheet: The Excel worksheet to be parsed.

        Returns:
            A list of ExcelTable objects representing the data tables.
        """
        bounds: DataRegion = self._find_true_data_bounds(
            sheet
        )  # The true data boundaries
        tables: list[ExcelTable] = []  # List to store found tables
        visited: set[tuple[int, int]] = set()  # Track already visited cells

        # Limit scan to actual data bounds
        for ri, row in enumerate(
            sheet.iter_rows(
                min_row=bounds.min_row,
                max_row=bounds.max_row,
                min_col=bounds.min_col,
                max_col=bounds.max_col,
                values_only=False,
            ),
            start=bounds.min_row - 1,
        ):
            for rj, cell in enumerate(row, start=bounds.min_col - 1):
                if cell.value is None or (ri, rj) in visited:
                    continue

                # If the cell starts a new table, find its bounds
                table_bounds, visited_cells = self._find_table_bounds(
                    sheet, ri, rj, bounds.max_row, bounds.max_col
                )

                visited.update(visited_cells)  # Mark these cells as visited
                tables.append(table_bounds)

        return tables

    def _find_table_bounds(
        self,
        sheet: Worksheet,
        start_row: int,
        start_col: int,
        max_row: int,
        max_col: int,
    ) -> tuple[ExcelTable, set[tuple[int, int]]]:
        """Determine the bounds of a compact rectangular table.

        Args:
            sheet: The Excel worksheet to be parsed.
            start_row: The row number of the starting cell.
            start_col: The column number of the starting cell.
            max_row: Maximum row boundary from true data bounds.
            max_col: Maximum column boundary from true data bounds.

        Returns:
            A tuple with an Excel table and a set of cell coordinates.
        """
        _log.debug("find_table_bounds")

        table_max_row = self._find_table_bottom(sheet, start_row, start_col, max_row)
        table_max_col = self._find_table_right(sheet, start_row, start_col, max_col)

        # Collect the data within the bounds
        data = []
        visited_cells: set[tuple[int, int]] = set()
        for ri, row in enumerate(
            sheet.iter_rows(
                min_row=start_row + 1,  # start_row is 0-based but iter_rows is 1-based
                max_row=table_max_row + 1,
                min_col=start_col + 1,
                max_col=table_max_col + 1,
                values_only=False,
            ),
            start_row,
        ):
            for rj, cell in enumerate(row, start_col):
                # Check if the cell belongs to a merged range
                row_span = 1
                col_span = 1

                for merged_range in sheet.merged_cells.ranges:
                    if (
                        merged_range.min_row <= ri + 1
                        and ri + 1 <= merged_range.max_row
                        and merged_range.min_col <= rj + 1
                        and rj + 1 <= merged_range.max_col
                    ):
                        row_span = merged_range.max_row - merged_range.min_row + 1
                        col_span = merged_range.max_col - merged_range.min_col + 1
                        break

                if (ri, rj) not in visited_cells:
                    data.append(
                        ExcelCell(
                            row=ri - start_row,
                            col=rj - start_col,
                            text=str(cell.value),
                            row_span=row_span,
                            col_span=col_span,
                        )
                    )

                    # Mark all cells in the span as visited
                    for span_row in range(ri, ri + row_span):
                        for span_col in range(rj, rj + col_span):
                            visited_cells.add((span_row, span_col))

        return (
            ExcelTable(
                anchor=(start_col, start_row),
                num_rows=table_max_row + 1 - start_row,
                num_cols=table_max_col + 1 - start_col,
                data=data,
            ),
            visited_cells,
        )

    def _find_table_bottom(
        self, sheet: Worksheet, start_row: int, start_col: int, max_row: int
    ) -> int:
        """Find the bottom boundary of a table.

        Args:
            sheet: The Excel worksheet to be parsed.
            start_row: The starting row of the table.
            start_col: The starting column of the table.
            max_row: Maximum row boundary from true data bounds.

        Returns:
            The row index representing the bottom boundary of the table.
        """
        table_max_row: int = start_row

        for ri, (cell,) in enumerate(
            sheet.iter_rows(
                min_row=start_row + 2,
                max_row=max_row,
                min_col=start_col + 1,
                max_col=start_col + 1,
                values_only=False,
            ),
            start_row + 1,
        ):
            # Check if the cell is part of a merged range
            merged_range = next(
                (mr for mr in sheet.merged_cells.ranges if cell.coordinate in mr),
                None,
            )

            if cell.value is None and not merged_range:
                break  # Stop if the cell is empty and not merged

            # Expand table_max_row to include the merged range if applicable
            if merged_range:
                table_max_row = max(table_max_row, merged_range.max_row - 1)
            else:
                table_max_row = ri

        return table_max_row

    def _find_table_right(
        self, sheet: Worksheet, start_row: int, start_col: int, max_col: int
    ) -> int:
        """Find the right boundary of a table.

        Args:
            sheet: The Excel worksheet to be parsed.
            start_row: The starting row of the table.
            start_col: The starting column of the table.
            max_col: The actual max column of the table.

        Returns:
            The column index representing the right boundary of the table."
        """
        table_max_col: int = start_col

        for rj, (cell,) in enumerate(
            sheet.iter_cols(
                min_row=start_row + 1,
                max_row=start_row + 1,
                min_col=start_col + 2,
                max_col=max_col,
                values_only=False,
            ),
            start_col + 1,
        ):
            # Check if the cell is part of a merged range
            merged_range = next(
                (mr for mr in sheet.merged_cells.ranges if cell.coordinate in mr),
                None,
            )

            if cell.value is None and not merged_range:
                break  # Stop if the cell is empty and not merged

            # Expand table_max_col to include the merged range if applicable
            if merged_range:
                table_max_col = max(table_max_col, merged_range.max_col - 1)
            else:
                table_max_col = rj

        return table_max_col

    def _find_images_in_sheet(
        self, doc: DoclingDocument, sheet: Worksheet
    ) -> DoclingDocument:
        """Find images in the Excel sheet and attach them to the DoclingDocument.

        Args:
            doc: The DoclingDocument to be updated.
            sheet: The Excel worksheet to be parsed.

        Returns:
            The updated DoclingDocument.
        """
        if self.workbook is not None:
            content_layer = self._get_sheet_content_layer(sheet)
            # Iterate over byte images in the sheet
            for item in sheet._images:  # type: ignore[attr-defined]
                try:
                    image: Image = cast(Image, item)
                    pil_image = PILImage.open(image.ref)  # type: ignore[arg-type]
                    page_no = self.workbook.index(sheet) + 1
                    anchor = (0, 0, 0, 0)
                    if isinstance(image.anchor, TwoCellAnchor):
                        anchor = (
                            image.anchor._from.col,
                            image.anchor._from.row,
                            image.anchor.to.col + 1,
                            image.anchor.to.row + 1,
                        )
                    doc.add_picture(
                        parent=self.parents[0],
                        image=ImageRef.from_pil(image=pil_image, dpi=72),
                        caption=None,
                        prov=ProvenanceItem(
                            page_no=page_no,
                            charspan=(0, 0),
                            bbox=BoundingBox.from_tuple(
                                anchor, origin=CoordOrigin.TOPLEFT
                            ),
                        ),
                        content_layer=content_layer,
                    )
                except Exception:
                    _log.error("could not extract the image from excel sheets")

        return doc

    @staticmethod
    def _find_page_size(
        doc: DoclingDocument, page_no: PositiveInt
    ) -> tuple[float, float]:
        left: float = -1.0
        top: float = -1.0
        right: float = -1.0
        bottom: float = -1.0
        for item, _ in doc.iterate_items(traverse_pictures=True, page_no=page_no):
            if not isinstance(item, DocItem):
                continue
            for provenance in item.prov:
                bbox = provenance.bbox
                left = min(left, bbox.l) if left != -1 else bbox.l
                right = max(right, bbox.r) if right != -1 else bbox.r
                top = min(top, bbox.t) if top != -1 else bbox.t
                bottom = max(bottom, bbox.b) if bottom != -1 else bbox.b

        return (right - left, bottom - top)

    @staticmethod
    def _get_sheet_content_layer(sheet: Worksheet) -> Optional[ContentLayer]:
        return (
            None
            if sheet.sheet_state == Worksheet.SHEETSTATE_VISIBLE
            else ContentLayer.INVISIBLE
        )
