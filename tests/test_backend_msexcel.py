import logging
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook

from docling.backend.msexcel_backend import MsExcelDocumentBackend
from docling.datamodel.base_models import InputFormat
from docling.datamodel.document import ConversionResult, DoclingDocument, InputDocument
from docling.document_converter import DocumentConverter

from .test_data_gen_flag import GEN_TEST_DATA
from .verify_utils import verify_document, verify_export

_log = logging.getLogger(__name__)

GENERATE = GEN_TEST_DATA


def get_excel_paths():
    # Define the directory you want to search
    directory = Path("./tests/data/xlsx/")

    # List all Excel files in the directory and its subdirectories
    excel_files = sorted(directory.rglob("*.xlsx")) + sorted(directory.rglob("*.xlsm"))
    return excel_files


def get_converter():
    converter = DocumentConverter(allowed_formats=[InputFormat.XLSX])

    return converter


@pytest.fixture(scope="module")
def documents() -> list[tuple[Path, DoclingDocument]]:
    documents: list[dict[Path, DoclingDocument]] = []

    excel_paths = get_excel_paths()
    converter = get_converter()

    for excel_path in excel_paths:
        _log.debug(f"converting {excel_path}")

        gt_path = (
            excel_path.parent.parent / "groundtruth" / "docling_v2" / excel_path.name
        )

        conv_result: ConversionResult = converter.convert(excel_path)

        doc: DoclingDocument = conv_result.document

        assert doc, f"Failed to convert document from file {gt_path}"
        documents.append((gt_path, doc))

    return documents


def test_e2e_excel_conversions(documents) -> None:
    for gt_path, doc in documents:
        pred_md: str = doc.export_to_markdown()
        assert verify_export(pred_md, str(gt_path) + ".md"), "export to md"

        pred_itxt: str = doc._export_to_indented_text(
            max_text_len=70, explicit_tables=False
        )
        assert verify_export(pred_itxt, str(gt_path) + ".itxt"), (
            "export to indented-text"
        )

        assert verify_document(doc, str(gt_path) + ".json", GENERATE), (
            "document document"
        )


def test_pages(documents) -> None:
    """Test the page count and page size of converted documents.

    Args:
        documents: The paths and converted documents.
    """
    # number of pages from the backend method
    path = next(item for item in get_excel_paths() if item.stem == "xlsx_01")
    in_doc = InputDocument(
        path_or_stream=path,
        format=InputFormat.XLSX,
        filename=path.stem,
        backend=MsExcelDocumentBackend,
    )
    backend = MsExcelDocumentBackend(in_doc=in_doc, path_or_stream=path)
    assert backend.page_count() == 4

    # number of pages from the converted document
    doc = next(item for path, item in documents if path.stem == "xlsx_01")
    assert len(doc.pages) == 4

    # page sizes as number of cells
    assert doc.pages.get(1).size.as_tuple() == (3.0, 7.0)
    assert doc.pages.get(2).size.as_tuple() == (9.0, 18.0)
    assert doc.pages.get(3).size.as_tuple() == (13.0, 36.0)
    assert doc.pages.get(4).size.as_tuple() == (0.0, 0.0)


def test_chartsheet(documents) -> None:
    """Test the conversion of Chartsheets.

    Args:
        documents: The paths and converted documents.
    """
    doc = next(item for path, item in documents if path.stem == "xlsx_03_chartsheet")
    assert len(doc.pages) == 2

    # Chartseet content is for now ignored
    assert doc.groups[1].name == "sheet: Duck Chart"
    assert doc.pages[2].size.height == 0
    assert doc.pages[2].size.width == 0


def test_chartsheet_data_values(documents) -> None:
    """Test that data values are extracted correctly from xlsx_03_chartsheet.

    This test verifies that calculated values (not formulas) are returned.
    The file contains duck observations with year 2024 having a total of 310 ducks.
    We need to verify that both 2024 and 310 appear in the parsed data.

    Args:
        documents: The paths and converted documents.
    """
    doc = next(item for path, item in documents if path.stem == "xlsx_03_chartsheet")

    # Find all tables
    tables = list(doc.tables)
    assert len(tables) > 0, "Should have at least one table"

    # Look for a table that has the year 2024 in it
    table_with_2024 = None
    row_index_of_2024 = None

    for table in tables:
        for cell in table.data.table_cells:
            if cell.text == "2024":
                table_with_2024 = table
                row_index_of_2024 = cell.start_row_offset_idx
                break
        if table_with_2024:
            break

    assert table_with_2024 is not None, "Should find a table containing year 2024"
    assert row_index_of_2024 is not None, "Should find row index for 2024"

    # Now verify that the value 310 exists in the document
    # (it may be in the same table or a different table due to how the parser splits tables)
    found_310 = False
    for table in tables:
        for cell in table.data.table_cells:
            if cell.text == "310":
                found_310 = True
                break
        if found_310:
            break

    assert found_310, "Should find the value 310 (total ducks for 2024) in the document"


def test_inflated_rows_handling(documents) -> None:
    """Test that files with inflated max_row are handled correctly.

    xlsx_04_inflated.xlsx has inflated max_row (1,048,496) but only 7 rows of actual data.
    This test verifies that our backend correctly identifies true data bounds.
    """
    # First, verify the file has inflated max_row using openpyxl directly
    path = next(item for item in get_excel_paths() if item.stem == "xlsx_04_inflated")

    wb = load_workbook(path)
    ws = wb.active
    reported_max_row = ws.max_row

    # Assert that openpyxl reports inflated max_row
    assert reported_max_row > 100000, (
        f"xlsx_04_inflated.xlsx should have inflated max_row (expected >100k, got {reported_max_row:,}). "
        f"This test file is designed to verify proper handling of Excel files with inflated row counts."
    )

    _log.info(
        f"xlsx_04_inflated.xlsx - Openpyxl reported max_row: {reported_max_row:,}"
    )

    # Now test that our backend handles it correctly
    in_doc = InputDocument(
        path_or_stream=path,
        format=InputFormat.XLSX,
        filename=path.stem,
        backend=MsExcelDocumentBackend,
    )
    backend = MsExcelDocumentBackend(in_doc=in_doc, path_or_stream=path)

    # Verify backend detects correct number of pages (should be 4, like test-01)
    page_count = backend.page_count()
    assert page_count == 4, (
        f"Backend should detect 4 pages (same as test-01), got {page_count}"
    )

    # Verify converted document has correct pages
    doc = next(item for path, item in documents if path.stem == "xlsx_04_inflated")
    assert len(doc.pages) == 4, f"Document should have 4 pages, got {len(doc.pages)}"

    # Verify page sizes match expected dimensions (same as test-01)
    # These should reflect actual data, not inflated row counts
    assert doc.pages.get(1).size.as_tuple() == (3.0, 7.0), (
        f"Page 1 should be 3x7 cells, got {doc.pages.get(1).size.as_tuple()}"
    )
    assert doc.pages.get(2).size.as_tuple() == (9.0, 18.0), (
        f"Page 2 should be 9x18 cells, got {doc.pages.get(2).size.as_tuple()}"
    )
    assert doc.pages.get(3).size.as_tuple() == (13.0, 36.0), (
        f"Page 3 should be 13x36 cells, got {doc.pages.get(3).size.as_tuple()}"
    )
    assert doc.pages.get(4).size.as_tuple() == (0.0, 0.0), (
        f"Page 4 should be 0x0 cells (empty), got {doc.pages.get(4).size.as_tuple()}"
    )

    _log.info(
        f"✓ Successfully handled inflated max_row: "
        f"reported {reported_max_row:,} rows, "
        f"correctly processed as {page_count} pages with proper dimensions"
    )


def test_bytesio_stream():
    """Test that Excel files can be loaded from BytesIO streams.

    This test verifies that the BytesIO code path in the backend is working correctly,
    ensuring that data_only=True is applied when loading workbooks from streams.
    """
    # Get a test Excel file
    path = next(item for item in get_excel_paths() if item.stem == "xlsx_01")

    # Load the file into a BytesIO stream
    buf = BytesIO(path.open("rb").read())

    # Create an InputDocument with the BytesIO stream
    in_doc = InputDocument(
        path_or_stream=buf,
        format=InputFormat.XLSX,
        filename=path.stem,
        backend=MsExcelDocumentBackend,
    )

    # Initialize the backend with the BytesIO stream
    backend = MsExcelDocumentBackend(in_doc=in_doc, path_or_stream=buf)

    # Verify the backend is valid
    assert backend.is_valid(), "Backend should be valid when loaded from BytesIO"

    # Verify page count matches expected value
    assert backend.page_count() == 4, "Should detect 4 pages from BytesIO stream"

    # Convert the document
    doc = backend.convert()

    # Verify the document was converted successfully
    assert doc is not None, "Document should be converted from BytesIO stream"
    assert len(doc.pages) == 4, "Document should have 4 pages"

    # Verify page sizes match expected dimensions
    assert doc.pages.get(1).size.as_tuple() == (3.0, 7.0)
    assert doc.pages.get(2).size.as_tuple() == (9.0, 18.0)
    assert doc.pages.get(3).size.as_tuple() == (13.0, 36.0)
    assert doc.pages.get(4).size.as_tuple() == (0.0, 0.0)
